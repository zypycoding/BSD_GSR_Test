# coding:utf-8
import json
import os
import shutil
import subprocess
import threading
import time
import matplotlib.pyplot as plt
import cv2
import pandas as pd
from matplotlib.font_manager import FontProperties
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from GSR_CV_BSD.common import commonscripts
from GSR_CV_BSD.common import surpose_algo

# 素材路径映射关系
test_conf_path = {
    "五月份素材": {
        "151": r"E:\Arcsoft\Zhangyun\欧标商用车\测试素材\5月份素材\151",
        "159": r"E:\Arcsoft\Zhangyun\欧标商用车\测试素材\5月份素材\159"},
    "七月份素材": r"\\hz-IOTFS\testlib_IOT\SmartCar\ADAS\Sample\ProjectSample\HIL_ADAS\ZY\欧标\测试素材\七月份素材",
    "九月份素材": r"\\hz-IOTFS\testlib_IOT\SmartCar\ADAS\Sample\ProjectSample\HIL_ADAS\ZY\欧标\测试素材\九月份素材",
    "金旅二号车素材": r"\\hz-IOTFS\testlib_IOT\SmartCar\ADAS\Sample\ProjectSample\HIL_ADAS\ZY\欧标\测试素材\金旅二号车素材",
    "金旅一号车素材": r"\\hz-IOTFS\testlib_IOT\SmartCar\ADAS\Sample\ProjectSample\HIL_ADAS\ZY\欧标\测试素材\金旅一号车素材",
    "扬州VAN素材": {"151": r"E:\Arcsoft\Zhangyun\欧标商用车\测试素材\扬州VAN素材\151",
                    "159": r"E:\Arcsoft\Zhangyun\欧标商用车\测试素材\扬州VAN素材\159"},
    "海格GSR素材": {
        "151": r"E:\Arcsoft\Zhangyun\欧标商用车\测试素材\海格GSR素材\151",
        "159": r"E:\Arcsoft\Zhangyun\欧标商用车\测试素材\海格GSR素材\159"
    },
}

# 算法txt与素材对应关系
algo_txt_path = {
    "五月份素材": r"D:\Arcsoft\ZhangYun\DevelopDocument\BSD_GSR\BSD_GSR_testbed\BSD_GSR\x64\Debug\input\videos\cal_test",
    "金旅二号车素材": r"D:\Arcsoft\ZhangYun\DevelopDocument\BSD_GSR\BSD_GSR_testbed\BSD_GSR\x64\Debug\input\videos\jinLv2",
    "金旅一号车素材": r"D:\Arcsoft\ZhangYun\DevelopDocument\BSD_GSR\BSD_GSR_testbed\BSD_GSR\x64\Debug\input\videos\jinLv1",
    "扬州VAN素材": r"D:\Arcsoft\ZhangYun\DevelopDocument\BSD_GSR\BSD_GSR_testbed\BSD_GSR\x64\Debug\input\videos\van_test",
    "海格GSR素材": r"D:\Arcsoft\ZhangYun\DevelopDocument\BSD_GSR\BSD_GSR_testbed\BSD_GSR\x64\Debug\input\videos\haige_test"
}

font = cv2.FONT_HERSHEY_SIMPLEX
# 控制图形属性
# 设置中文字体
plt.rcParams["font.sans-serif"] = "SimHei"
plt.rcParams["axes.unicode_minus"] = False


class Tool:
    def __init__(self):
        self.select_shot_l = None
        self.select_input_shot_type = ""
        self.select_input_case_type = ""
        self.algo_result_path = None
        self.log_path = None
        self.select_output_source = r""
        self.select_input_source_type = None
        self.select_input_rule_type = None
        self.src_test_conf_path = None
        self.src_algo_txt_path = None
        self.custom_result = None
        self.rect_type = None

    def pre_test_bed(self, select_test_conf_path):
        for root, dirs, files in os.walk(select_test_conf_path):
            for file in files:
                if file.endswith(".mp4"):
                    for shot in self.select_shot_l:
                        if shot in file:
                            if (
                                    self.select_input_source_type == "五月份素材" or self.select_input_source_type == "扬州VAN素材") and (
                                    self.select_input_rule_type == "151" or (
                                    self.select_input_rule_type == "159" and self.select_input_case_type == "驶离")):
                                cal_path = os.path.join(root, "..", "..", "config", file.replace(".mp4", ".cal"))
                                shutil.copy2(cal_path, self.src_algo_txt_path)
                            mp4_path = os.path.join(root, file)
                            shutil.copy2(mp4_path, self.src_algo_txt_path)

        config_json = json.loads(
            open(r"D:\Arcsoft\ZhangYun\DevelopDocument\BSD_GSR\BSD_GSR_testbed\BSD_GSR\x64\Debug\config.json",
                 'r').read())
        replace_value = self.src_algo_txt_path.split("\\")[-1]
        config_json["select"] = replace_value
        with open(r"D:\Arcsoft\ZhangYun\DevelopDocument\BSD_GSR\BSD_GSR_testbed\BSD_GSR\x64\Debug\config.json",
                  'w') as fw:
            json_str = json.dumps(config_json, indent=4)
            fw.write(json_str)
            fw.write('\n')
        ex1 = commonscripts.Func()
        ex1.run_func(self.src_algo_txt_path)

    def run_test_bed(self):
        try:
            # 生成BSD_GSR 可执行文件
            result = subprocess.run(
                ['devenv', 'D:\\Arcsoft\\ZhangYun\\DevelopDocument\\BSD_GSR\\BSD_GSR_testbed\\BSD_GSR\\BSD.sln',
                 '/Build', 'Debug|x64'], check=True, capture_output=True, text=True)
        except subprocess.CalledProcessError as e:
            print(f"命令执行失败，错误信息：{e}")
        else:
            print("生成 BSD_GSR.exe 成功")
        # 获取当前工作目录
        current_dir = os.getcwd()
        print("当前工作目录：", current_dir)
        # 切换到指定工作目录
        new_dir = "D:\\Arcsoft\\ZhangYun\\DevelopDocument\\BSD_GSR\\BSD_GSR_testbed\\BSD_GSR\\x64\\Debug"
        os.chdir(new_dir)
        # 再次获取当前工作目录，确认是否切换成功
        current_dir = os.getcwd()
        print("切换后的工作目录：", current_dir)
        try:
            result = subprocess.run(['BSD_GSR.exe'], capture_output=True, text=True)
            # print("Return code:", result.returncode)
            # print("Output:", result.stdout)
            with open(os.path.join(self.log_path, f"bsd_gsr_{self.select_input_source_type}_log.txt"), 'a+') as f:
                f.write(result.stdout)
        except FileNotFoundError as e:
            print("文件未找到，请检查文件路径是否正确。")
        else:
            print(f"testbed运行log写入 bsd_gsr_{self.select_input_source_type}_log.txt 完成")

    def copy_test_con(self, select_test_conf_path):
        for root, dirs, files in os.walk(select_test_conf_path):
            for file in files:
                if file.endswith(".mp4"):
                    for shot in self.select_shot_l:
                        if shot in file:
                            if (
                                    self.select_input_source_type == "五月份素材" or self.select_input_source_type == "扬州VAN素材") and (
                                    self.select_input_rule_type == "151" or (
                                    self.select_input_rule_type == "159" and self.select_input_case_type == "驶离")):
                                cal_path = os.path.join(root, "..", "..", "config", file.replace(".mp4", ".cal"))
                                print(cal_path)
                                shutil.copy2(cal_path, self.algo_result_path)
                            # base_file = os.path.splitext(file)[0]
                            mp4_path = os.path.join(root, file)
                            algo_file_path = file.replace(".mp4", ".txt")
                            algo_path = os.path.join(self.src_algo_txt_path, algo_file_path)
                            rt_path = os.path.join(root, "..", "..", "config", file.replace(".mp4", "_rt.txt"))
                            pts_path = os.path.join(root, "..", "..", "config", file.replace(".mp4", ".log"))
                            custom_path = os.path.join(root, "..", "..", "config", file.replace(".mp4", "_custom.txt"))
                            shutil.copy2(mp4_path, self.algo_result_path)
                            shutil.copy2(rt_path, self.algo_result_path)
                            shutil.copy2(pts_path, self.algo_result_path)
                            shutil.copy2(custom_path, self.algo_result_path)
                            shutil.copy2(algo_path, self.algo_result_path)

    def pre_pose(self):
        # selec_input_source = input("请输入你要处理的素材:").strip()
        self.select_input_source_type = "扬州VAN素材"
        self.select_input_rule_type = "151"
        self.rect_type = 'bike'
        # 输出路径
        self.select_output_source = r"E:\Arcsoft\Zhangyun\测试记录\test"
        self.log_path = self.select_output_source + '\\output\\log'
        self.algo_result_path = self.select_output_source + '\\output\\result'
        # 选择需要执行的素材
        if os.path.exists(self.log_path):
            shutil.rmtree(self.log_path)
        os.makedirs(self.log_path)
        if os.path.exists(self.algo_result_path):
            shutil.rmtree(self.algo_result_path)
        os.makedirs(self.algo_result_path)
        self.src_test_conf_path = test_conf_path[self.select_input_source_type][self.select_input_rule_type]
        self.src_algo_txt_path = algo_txt_path[self.select_input_source_type]
        if os.path.exists(self.src_algo_txt_path):
            shutil.rmtree(self.src_algo_txt_path)
        os.makedirs(self.src_algo_txt_path)
        if self.select_input_rule_type == "151":
            print("请确认是否需要测试151全部素材 case1到case7：输入 Y or N")
            confirm_flag = "N"
            if confirm_flag == "Y":
                self.select_input_case_type = "case1-7"
            elif confirm_flag == "N":
                print("请输入需要测试的素材（请用空格做分割符):\n示例1：case1 case2 case3\n")
                self.select_input_case_type = "case1"
                print(
                    "请输入需要测试的镜头类别（请用空格做分割符）\nA 是后平视，B 是前鱼眼， C是右鱼眼:\n示例1：A B C\n示例2：A B\n")
                self.select_input_shot_type = "A"
        elif self.select_input_rule_type == "159":
            print("请确认是否需要测试159全部素材 驶离 + 横穿：输入 Y or N")
            confirm_flag = "N"
            if confirm_flag == "Y":
                self.select_input_case_type = "驶离and横穿"
            elif confirm_flag == "N":
                print("请输入需要测试的素材（请用空格做分割符):\n示例1：驶离\n示例2：横穿")
                self.select_input_case_type = "横穿"
                print(
                    "请输入需要测试的镜头类别，若不输入默认为前鱼眼（请用空格做分割符）\nA 是后平视，B 是前鱼眼， C 是右鱼眼:\n示例1：A B C\n示例2：A B\n")
                self.select_input_shot_type = "B"
        select_case_path_l = self.select_input_case_type.strip().split(" ")
        self.select_shot_l = self.select_input_shot_type.strip().split(" ")
        for select_case_path in select_case_path_l:
            select_test_conf_path = self.src_test_conf_path + "\\" + "video" + "\\" + select_case_path
            self.pre_test_bed(select_test_conf_path)
        self.run_test_bed()
        for select_case_path in select_case_path_l:
            select_test_conf_path = self.src_test_conf_path + "\\" + "video" + "\\" + select_case_path
            self.copy_test_con(select_test_conf_path)

    def get_files_list(self, path, keywords):
        files_list = []
        for root, dirs, files in os.walk(path):
            for file in files:
                if file.endswith(keywords):
                    file_path = os.path.join(root, file)
                    files_list.append(file_path)
        return files_list

    def final_order_result(self, mp4_path, pts_path, algo_path, rt_result_path):
        cap = cv2.VideoCapture(mp4_path)  # 读取视频文件
        # 获取视频帧率
        fps = cap.get(cv2.CAP_PROP_FPS)
        # 获取视频宽度和高度
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        # 保存回放视频
        fourcc = cv2.VideoWriter_fourcc(*'MJPG')  # AVI格式
        result_video = cv2.VideoWriter(mp4_path.replace('.mp4', '_result.avi'), fourcc, fps, (width, height))

        cal_list = open(mp4_path.replace('mp4', 'cal'), 'r').readlines()  # 整个cal的list

        # 结果保存为excel表格
        workbook = Workbook()  # 在本地创建Excel工作簿
        # worksheet = workbook.active
        w1 = workbook.create_sheet("test", 0)  # 第一个sheet
        # 首行，名称，范围0~40m,步长2m
        default_name = ['帧号', '时间戳', '算法横向测距', '算法纵向测距', '算法横向测速', '算法纵向测速',
                        '真值横向测距', '真值纵向测距', '真值横向测速', '真值纵向测速', '横向测距误差', '纵向测距误差',
                        '横向测速误差', '纵向测速误差', '算法运动状态', '真值运动状态']
        w1.append(default_name)

        algo_result = open(algo_path, 'r').readlines()
        pts_result = open(pts_path, 'r').readlines()
        rt_result = open(rt_result_path, 'r').readlines()
        frame_index = 0

        # 绘制算法、真值等信息到结果视频中
        while cap.isOpened():
            ret, frame = cap.read()  # 获取一帧视频图像
            if not ret:
                # print(' can not receive frame,exiting ...')
                commonscripts.LogColor.info('{} 算法,真值 绘制完成'.format(mp4_path))
                break
                # continue
            algo_result_index = eval(algo_result[frame_index]).get('mres_person', [])

            iou_map = dict()
            iou_temp = 0
            for algo_result_iou in algo_result_index:
                cal_list_e = json.loads(cal_list[frame_index]).get('rect', {'left': 0, 'top': 0, 'right': 0,
                                                                            'bottom': 0, })  # 由于算法误检，真值为空时，就检测到
                mres_rect = algo_result_iou.get('mres_rect')
                mres_rect_set = mres_rect.split(',')
                mres_rect_set_int = tuple(map(int, mres_rect_set))
                cal_list_set = (
                    int(cal_list_e['left']), int(cal_list_e['top']), int(cal_list_e['right']),
                    int(cal_list_e['bottom']))
                IOU = commonscripts.RectData(mres_rect_set_int, cal_list_set).iou()[0]
                iou_temp = max(IOU, iou_temp)
                iou_map[IOU] = algo_result_iou

            if len(iou_map) == 0:
                LatRange_e = eval(rt_result[frame_index]).get('LatRange', 0)
                LongRange_e = eval(rt_result[frame_index]).get('LongRange', 0)
                LatVel_e = eval(rt_result[frame_index]).get('LatVel', 0)
                LongVel_e = eval(rt_result[frame_index]).get('LongVel', 0)
                if frame_index in range(self.custom_result['mv_frame'][0], self.custom_result['mv_frame'][1] + 1):
                    Mv_status_e = 1
                else:
                    Mv_status_e = 0

                # VAN素材需要对真值结果做前轴的补偿
                if self.select_input_source_type == "扬州VAN素材":
                    if self.select_input_rule_type == "151":
                        LongRange_e = LongRange_e + 3.75
                        # 151补偿自行车长度
                        if self.rect_type == 'bike':
                            if '_A' in mp4_path:
                                LongRange_e = LongRange_e + 1.01
                            elif '_B' in mp4_path:
                                LongRange_e = LongRange_e - 1.01
                            elif '_C' in mp4_path:
                                LongRange_e = LongRange_e - 1.01
                    # van素材159驶离素材真值纵向补偿2.4m
                    elif self.select_input_rule_type == "159" and self.select_input_case_type == "驶离":
                        LongRange_e = LongRange_e + 2.4
                        # 159驶离补偿自行车长度
                        if self.rect_type == 'bike':
                            if '_A' in mp4_path:
                                LongRange_e = LongRange_e + 1.01
                            elif '_B' in mp4_path:
                                LongRange_e = LongRange_e - 1.01
                            elif '_C' in mp4_path:
                                LongRange_e = LongRange_e - 1.01
                    # 海格素材不做任何补偿
                elif self.select_input_source_type == "海格GSR素材":
                    pass
                elif self.select_input_source_type == "五月份素材":
                    # 补偿自行车长度
                    if '_A' in mp4_path:
                        LongRange_e = LongRange_e + 1.01
                    elif '_B' in mp4_path:
                        LongRange_e = LongRange_e - 1.01
                    elif '_C' in mp4_path:
                        LongRange_e = LongRange_e - 1.01

                info_algo_e = f'algo  LatDis:{0} LongDis:{0} LatVel:{0} LongVel:{0}'
                info_rt_e = f'RT  LatDis:{round(LatRange_e, 2)} LongDis:{round(LongRange_e, 2)} LatVel:{round(LatVel_e, 2)} LongVel:{round(LongVel_e, 2)}'
                info_mv_e = f'algo  mv_state:{0}   RT  mv_state:{Mv_status_e}'
                info_frame_id = f'frame_id:{frame_index}'
                cv2.putText(frame, info_algo_e, (10, 50), font, 1.2, (255, 255, 0), 2)
                cv2.putText(frame, info_rt_e, (10, 100), font, 1.2, (255, 255, 0), 2)
                cv2.putText(frame, info_mv_e, (10, 150), font, 1.2, (255, 255, 0), 2)
                cv2.putText(frame, info_frame_id, (10, 200), font, 1.2, (255, 255, 0), 2)
                result_video.write(frame)
                pts_e = eval(pts_result[frame_index]).get('pts')
                info_list_e = [frame_index, str(pts_e), (0), (0),
                               (0), (0), (LatRange_e), (LongRange_e),
                               (LatVel_e), (LongVel_e)]

                LatDis_offset = 0
                LongDis_offset = 0
                LatVel_offset = 0
                LongVel_offset = 0
                mv_state_e = 0
                info_list_e.extend([(LatDis_offset), (LongDis_offset), (LatVel_offset), (LongVel_offset)])
                info_list_e.append(mv_state_e)
                info_list_e.append(Mv_status_e)
                if frame_index in range(self.custom_result['good_frame'][0], self.custom_result['good_frame'][1] + 1):
                    w1.append(info_list_e)
            else:
                for iou_index in iou_map:
                    if iou_index == iou_temp:
                        algo_result_index_e = iou_map[iou_index]
                        mres_rect1 = algo_result_index_e.get('mres_rect')
                        mres_rect_point = mres_rect1.split(',')
                        point_l = (int(mres_rect_point[0]), int(mres_rect_point[1]))
                        point_r = (int(mres_rect_point[2]), int(mres_rect_point[3]))
                        cv2.rectangle(frame, point_l, point_r, (0, 255, 0), 2)
                        mres_id = algo_result_index_e.get('mres_id')
                        # print(mp4_path,mres_id)
                        cv2.putText(frame, str(mres_id), point_l, font, 1.2, (255, 255, 255), 2)
                        pts_e = eval(pts_result[frame_index]).get('pts')
                        try:
                            f32LatDistance_e = algo_result_index_e.get('f32LatDistance') + 1
                            f32LongDistance_e = algo_result_index_e.get('f32LongDistance')
                            f32AbsoluteLatVelocity_e = algo_result_index_e.get('f32AbsoluteLatVelocity')
                            f32AbsoluteLongVelocity_e = algo_result_index_e.get('f32AbsoluteLongVelocity')
                            mv_state_e = algo_result_index_e.get('mv_state')
                        except:
                            f32LatDistance_e = 0
                            f32LongDistance_e = 0
                            f32AbsoluteLatVelocity_e = 0
                            f32AbsoluteLongVelocity_e = 0
                            mv_state_e = 0
                        LatRange_e = eval(rt_result[frame_index]).get('LatRange', 0) + 1
                        LongRange_e = eval(rt_result[frame_index]).get('LongRange', 0)
                        LatVel_e = eval(rt_result[frame_index]).get('LatVel', 0)
                        LongVel_e = eval(rt_result[frame_index]).get('LongVel', 0)
                        if frame_index in range(self.custom_result['mv_frame'][0],
                                                self.custom_result['mv_frame'][1] + 1):
                            Mv_status_e = 1
                        else:
                            Mv_status_e = 0

                        # VAN素材需要对真值结果做前轴的补偿
                        if self.select_input_source_type == "扬州VAN素材":
                            if self.select_input_rule_type == "151":
                                LongRange_e = LongRange_e + 3.75
                                # 151补偿自行车长度
                                if self.rect_type == 'bike':
                                    if '_A' in mp4_path:
                                        LongRange_e = LongRange_e + 1.01
                                    elif '_B' in mp4_path:
                                        LongRange_e = LongRange_e - 1.01
                                    elif '_C' in mp4_path:
                                        LongRange_e = LongRange_e - 1.01
                            # van素材159驶离素材真值纵向补偿2.4m
                            elif self.select_input_rule_type == "159" and self.select_input_case_type == "驶离":
                                LongRange_e = LongRange_e + 2.4
                                # 159驶离补偿自行车长度
                                if self.rect_type == 'bike':
                                    if '_A' in mp4_path:
                                        LongRange_e = LongRange_e + 1.01
                                    elif '_B' in mp4_path:
                                        LongRange_e = LongRange_e - 1.01
                                    elif '_C' in mp4_path:
                                        LongRange_e = LongRange_e - 1.01
                        # 海格素材不做任何补偿
                        elif self.select_input_source_type == "海格GSR素材":
                            pass
                        elif self.select_input_source_type == "五月份素材":
                            # 补偿自行车长度
                            if '_A' in mp4_path:
                                LongRange_e = LongRange_e + 1.01
                            elif '_B' in mp4_path:
                                LongRange_e = LongRange_e - 1.01
                            elif '_C' in mp4_path:
                                LongRange_e = LongRange_e - 1.01

                        info_algo_e = f'algo  LatDis:{round(f32LatDistance_e, 2)} LongDis:{round(f32LongDistance_e, 2)} LatVel:{round(f32AbsoluteLatVelocity_e, 2)} LongVel:{round(f32AbsoluteLongVelocity_e, 2)}'
                        info_rt_e = f'RT  LatDis:{round(LatRange_e, 2)} LongDis:{round(LongRange_e, 2)} LatVel:{round(LatVel_e, 2)} LongVel:{round(LongVel_e, 2)}'
                        info_mv_e = f'algo  mv_state:{mv_state_e}   RT  mv_state:{Mv_status_e}'
                        info_frame_id = f'frame_id:{frame_index}'
                        cv2.putText(frame, info_algo_e, (10, 50), font, 1.2, (255, 255, 0), 2)
                        cv2.putText(frame, info_rt_e, (10, 100), font, 1.2, (255, 255, 0), 2)
                        cv2.putText(frame, info_mv_e, (10, 150), font, 1.2, (255, 255, 0), 2)
                        cv2.putText(frame, info_frame_id, (10, 200), font, 1.2, (255, 255, 0), 2)

                        result_video.write(frame)

                        info_list_e = [frame_index, str(pts_e), (f32LatDistance_e), (f32LongDistance_e),
                                       (f32AbsoluteLatVelocity_e), (f32AbsoluteLongVelocity_e), (LatRange_e),
                                       (LongRange_e),
                                       (LatVel_e), (LongVel_e)]

                        if f32LatDistance_e == 0:
                            LatDis_offset = 0
                        else:
                            LatDis_offset = f32LatDistance_e - LatRange_e
                        if f32LongDistance_e == 0:
                            LongDis_offset = 0
                        else:
                            LongDis_offset = f32LongDistance_e - LongRange_e
                        if f32AbsoluteLatVelocity_e == 0:
                            LatVel_offset = 0
                        else:
                            LatVel_offset = f32AbsoluteLatVelocity_e - LatVel_e
                        if f32AbsoluteLongVelocity_e == 0:
                            LongVel_offset = 0
                        else:
                            LongVel_offset = f32AbsoluteLongVelocity_e - LongVel_e
                        # info_list_e.extend([abs(LatDis_offset),abs(LongDis_offset),abs(LatVel_offset),abs(LongVel_offset)])
                        info_list_e.extend([(LatDis_offset), (LongDis_offset), (LatVel_offset), (LongVel_offset)])
                        info_list_e.append(mv_state_e)
                        info_list_e.append(Mv_status_e)
                        if frame_index in range(self.custom_result['good_frame'][0],
                                                self.custom_result['good_frame'][1] + 1):
                            w1.append(info_list_e)
            frame_index = frame_index + 1

        cap.release()
        result_video.release()
        xlsx_save_file = mp4_path.replace('.mp4', '.xlsx')
        workbook.save(filename=xlsx_save_file)
        commonscripts.LogColor.info(f'{xlsx_save_file} 写入完成')

    def count_interval_result_new(self, df, w1, start, end):
        long_RT = df[((df[f"真值纵向测距"] >= start) & (df[f"真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "真值纵向测距"]
        lat_RT = df[((df[f"真值纵向测距"] >= start) & (df[f"真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "真值横向测距"]
        longvel_RT = df[((df[f"真值纵向测距"] >= start) & (df[f"真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "真值纵向测速"]
        latvel_RT = df[((df[f"真值纵向测距"] >= start) & (df[f"真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "真值横向测速"]
        long_error = df[((df["真值纵向测距"] >= start) & (df["真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "纵向测距误差"]
        lat_error = df[((df["真值纵向测距"] >= start) & (df["真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "横向测距误差"]
        longvel_error = df[((df["真值纵向测距"] >= start) & (df["真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "纵向测速误差"]
        latvel_error = df[((df["真值纵向测距"] >= start) & (df["真值纵向测距"] < end)) & (df["算法纵向测距"] != 0)][
            "横向测速误差"]
        # 计算平均值
        long_RT_avg = round(abs(long_RT).mean(), 4)  # 保留4位小数
        lat_RT_avg = round(abs(lat_RT).mean(), 4)
        longvel_RT_avg = round(abs(longvel_RT).mean(), 4)
        latvel_RT_avg = round(abs(latvel_RT).mean(), 4)
        long_error_avg = round(abs(long_error).mean(), 4)
        lat_error_avg = round(abs(lat_error).mean(), 4)
        longvel_error_avg = round(abs(longvel_error).mean(), 4)
        latvel_error_avg = round(abs(latvel_error).mean(), 4)

        # 误差占比(误差精度%）
        long_error_per = "{:.2f}%".format(long_error_avg / long_RT_avg * 100)
        lat_error_per = "{:.2f}%".format(lat_error_avg / lat_RT_avg * 100)
        longvel_error_per = "{:.2f}%".format(longvel_error_avg / longvel_RT_avg * 100)
        latvel_error_per = "{:.2f}%".format(latvel_error_avg / latvel_RT_avg * 100)

        # 最大误差
        long_error_max = round(abs(long_error).max(), 4)
        lat_error_max = round(abs(lat_error).max(), 4)
        longvel_error_max = round(abs(longvel_error).max(), 4)
        latvel_error_max = round(abs(latvel_error).max(), 4)

        # 获取最大误差对应帧号、真值纵向距离
        try:
            long_error_max_index = long_error.abs().idxmax() + self.custom_result['good_frame'][0]
            selected_row = df[df['帧号'] == long_error_max_index]
            long_error_max_index_RT = selected_row.loc[selected_row.index[0], '真值纵向测距']

            lat_error_max_index = lat_error.abs().idxmax() + self.custom_result['good_frame'][0]
            selected_row = df[df['帧号'] == lat_error_max_index]
            lat_error_max_index_RT = selected_row.loc[selected_row.index[0], '真值纵向测距']

            longvel_error_max_index = longvel_error.abs().idxmax() + self.custom_result['good_frame'][0]
            selected_row = df[df['帧号'] == longvel_error_max_index]
            longvel_error_max_index_RT = selected_row.loc[selected_row.index[0], '真值纵向测距']

            latvel_error_max_index = latvel_error.abs().idxmax() + self.custom_result['good_frame'][0]
            selected_row = df[df['帧号'] == latvel_error_max_index]
            latvel_error_max_index_RT = selected_row.loc[selected_row.index[0], '真值纵向测距']
        except:
            long_error_max_index = 'nan'
            long_error_max_index_RT = 'nan'
            lat_error_max_index = 'nan'
            lat_error_max_index_RT = 'nan'
            longvel_error_max_index = 'nan'
            longvel_error_max_index_RT = 'nan'
            latvel_error_max_index = 'nan'
            latvel_error_max_index_RT = 'nan'

        # avg-test表的9个列
        list_result = [f'[{start}, {end}]', f'{lat_RT_avg}', f'{long_RT_avg}', f'{latvel_RT_avg}', f'{longvel_RT_avg}',
                       f'{lat_error_per}({lat_error_avg}_{lat_error_max}_{lat_error_max_index}_{lat_error_max_index_RT})',
                       f'{long_error_per}({long_error_avg}_{long_error_max}_{long_error_max_index}_{long_error_max_index_RT})',
                       f'{latvel_error_per}({latvel_error_avg}_{latvel_error_max}_{latvel_error_max_index}_{latvel_error_max_index_RT})',
                       f'{longvel_error_per}({longvel_error_avg}_{longvel_error_max}_{longvel_error_max_index}_{longvel_error_max_index_RT})', ]
        w1.append(list_result)

    def get_ave_by_interval_new(self, xlxs_path):
        # 根据RT的纵向区间，分段计算均值和最大值
        df = pd.read_excel(xlxs_path, sheet_name='test')
        workbook = load_workbook(xlxs_path)
        # worksheet = workbook.active
        w1 = workbook.create_sheet("avg-test")
        default_name = ['纵向距离区间', '真值横向距离平均值', '真值纵向距离平均值', '真值横向速度平均值',
                        '真值纵向速度平均值',
                        '横向测距误差占比(误差平均值_最大误差_帧号_真值纵向距离)',
                        '纵向测距误差占比(误差平均值_最大误差_帧号_真值纵向距离)',
                        '横向测速误差占比(误差平均值_最大误差_帧号_真值纵向距离)',
                        '纵向测速误差占比(误差平均值_最大误差_帧号_真值纵向距离)']
        w1.append(default_name)
        if self.select_input_rule_type == "159":
            interval_list = [-3.01, -1.01, 1.01, 3.01]  # ACB
        else:
            interval_list = [-40, -30, -20, -11.94, -10, -5, 0, 3, 5, 10, 15, 20, 30, 40]  # ACB

        for index in range(len(interval_list) - 1):
            self.count_interval_result_new(df, w1, interval_list[index], interval_list[index + 1])
        self.count_interval_result_new(df, w1, interval_list[0], interval_list[-1])
        workbook.save(filename=xlxs_path)

    def draw_pic_result(self, mp4_path):
        df = pd.read_excel(mp4_path.replace('.mp4', '.xlsx'), sheet_name='test')
        x_axis = df["帧号"]
        y_algo_LatDis = df["算法横向测距"]
        y_RT_LatDis = df["真值横向测距"]
        y_LatDis_offset = df["横向测距误差"]

        y_algo_LongDis = df["算法纵向测距"]
        y_RT_LongDis = df["真值纵向测距"]
        y_LongDis_offset = df["纵向测距误差"]

        y_algo_LatVel = df["算法横向测速"]
        y_RT_LatVel = df["真值横向测速"]
        y_LatVel_offset = df["横向测速误差"]

        y_algo_LongVel = df["算法纵向测速"]
        y_RT_LongVel = df["真值纵向测速"]
        y_LongVel_offset = df["纵向测速误差"]

        fig, ax = plt.subplots(figsize=(12, 10))  # 创建更大的图形
        # 绘制图表
        ax.plot(x_axis, y_algo_LatDis, label="算法横向测距", linewidth=1, color='red')
        ax.plot(x_axis, y_RT_LatDis, label="真值横向测距", linewidth=1, color='green')
        ax.grid(True)
        # 添加图例
        ax.legend()
        ax.set_title(f'横向测距')
        fig.tight_layout()
        fig.savefig(mp4_path.replace('.mp4', '_横向测距.jpg'))  # 对比图
        ax.clear()
        ax.plot(x_axis, y_LatDis_offset, label="横向测距误差", linewidth=1, color='red')
        ax.grid(True)
        ax.set_title(f'横向测距误差')
        fig.tight_layout()
        # 去除零值，并求绝对值
        non_zero_values = abs(df[df["横向测距误差"] != 0]["横向测距误差"])
        # 计算平均值
        average = round(non_zero_values.mean(), 4)
        fig.savefig(mp4_path.replace('.mp4', f'_横向测距误差.jpg'))
        ax.clear()

        ax.plot(x_axis, y_algo_LongDis, label="算法纵向测距", linewidth=1, color='red')
        ax.plot(x_axis, y_RT_LongDis, label="真值纵向测距", linewidth=1, color='green')
        ax.grid(True)
        # 添加图例
        ax.legend()
        ax.set_title(f'纵向测距')
        fig.tight_layout()
        fig.savefig(mp4_path.replace('.mp4', '_纵向测距.jpg'))
        ax.clear()
        ax.plot(x_axis, y_LongDis_offset, label="纵向测距误差", linewidth=1, color='red')
        ax.grid(True)
        ax.set_title(f'纵向测距误差')
        fig.tight_layout()
        # 去除零值，并求绝对值
        non_zero_values = abs(df[df["纵向测距误差"] != 0]["纵向测距误差"])
        # 计算平均值
        average = round(non_zero_values.mean(), 4)
        fig.savefig(mp4_path.replace('.mp4', f'_纵向测距误差.jpg'))
        ax.clear()

        ax.plot(x_axis, y_algo_LatVel, label="算法横向测速", linewidth=1, color='red')
        ax.plot(x_axis, y_RT_LatVel, label="真值横向测速", linewidth=1, color='green')
        ax.grid(True)
        # 添加图例
        ax.legend()
        ax.set_title(f'横向测速')
        fig.tight_layout()
        fig.savefig(mp4_path.replace('.mp4', '_横向测速.jpg'))
        ax.clear()
        ax.plot(x_axis, y_LatVel_offset, label="横向测速误差", linewidth=1, color='red')
        ax.grid(True)
        ax.set_title(f'横向测速误差')
        fig.tight_layout()
        # 去除零值，并求绝对值
        non_zero_values = abs(df[df["横向测速误差"] != 0]["横向测速误差"])
        # 计算平均值
        average = round(non_zero_values.mean(), 4)
        fig.savefig(mp4_path.replace('.mp4', f'_横向测速误差.jpg'))
        ax.clear()

        ax.plot(x_axis, y_algo_LongVel, label="算法纵向测速", linewidth=1, color='red')
        ax.plot(x_axis, y_RT_LongVel, label="真值纵向测速", linewidth=1, color='green')
        ax.grid(True)
        # 添加图例
        ax.legend()
        ax.set_title(f'纵向测速')
        fig.tight_layout()
        fig.savefig(mp4_path.replace('.mp4', '_纵向测速.jpg'))
        ax.clear()
        ax.plot(x_axis, y_LongVel_offset, label="纵向测速误差", linewidth=1, color='red')
        ax.grid(True)
        ax.set_title(f'纵向测速误差')
        fig.tight_layout()
        # 去除零值，并求绝对值
        non_zero_values = abs(df[df["纵向测速误差"] != 0]["纵向测速误差"])
        # 计算平均值
        average = round(non_zero_values.mean(), 4)
        fig.savefig(mp4_path.replace('.mp4', f'_纵向测速误差.jpg'))
        ax.clear()

    def run(self):
        self.pre_pose()
        files_list = self.get_files_list(self.algo_result_path, ".mp4")
        for video_path in files_list:
            mp4_path = video_path
            rt_path = mp4_path.replace(".mp4", "_rt.txt")
            pts_path = mp4_path.replace(".mp4", ".log")
            algo_path = mp4_path.replace(".mp4", ".txt")
            custom_path = mp4_path.replace(".mp4", "_custom.txt")
            print(
                "请选择视频叠加模式：（不输入默认使用模式1生成完整测试结果）\n输入 1 叠加算法+真值，生成完整测试结果\n输入 2 只叠加算法结果")
            surpose_mode = "2"
            if surpose_mode == "1":
                self.custom_result = eval(open(custom_path, 'r').read())
                self.final_order_result(mp4_path, pts_path, algo_path, rt_path)
                self.get_ave_by_interval_new(mp4_path.replace('.mp4', '.xlsx'))
                self.draw_pic_result(mp4_path)
            elif surpose_mode == "2":
                surpose_algo.surpose_algo_result(video_path)


if __name__ == "__main__":
    ex = Tool()
    ex.run()
